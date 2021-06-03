import pandas as pd
import numpy as np
#from scipy.spatial import distance
from math import dist
import os
import csv
import openpyxl
from sklearn.metrics.cluster import adjusted_rand_score

# https://github.com/cran/dendextend/blob/master/R/find_k.R
# https://cran.r-project.org/web/packages/fpc/fpc.pdf

# scipy.spatial.distance.euclidean(A, B)
# dist([1, 0, 0], [0, 1, 0])

labels_true_gt=np.genfromtxt("C:/Users/dey.sn/Downloads/temp/haraal/haraal_labels_gt.csv", delimiter=',')
print(labels_true_gt.shape[0])
print(labels_true_gt)
#column = nArr2D[:, 1]
#output_labels = np.genfromtxt('C:/Users/dey.sn/Downloads/work/output/har_k6/Labels_har_k6_kmpp,cutoff,90,k6.csv', delimiter=',')
'''
output_labels_col1=output_labels[:,0]
print(output_labels.shape[1])
print(output_labels_col1)
for cols in range(output_labels.shape[1]):
    print(adjusted_rand_score(labels_true_gt,output_labels[:,cols]))

'''
# This is the path where you want to search
path = r'C:/Users/dey.sn/Downloads/work/output/haraal_k6/'
# this is the extension you want to detect
extension = '.csv'
substring="Labels"
count=0
wb=openpyxl.Workbook()
sheet=wb.active
sheet.title= 'haraal_ari'
for root, dirs_list, files_list in os.walk(path):
    for file_name in files_list:
       if os.path.splitext(file_name)[-1] == extension:
          file_name_path = os.path.join(root, file_name)
          print(file_name)
          print(file_name_path)   # This is the full path of the filter file
          try:
            index=file_name.index(substring)
 #          print(index)
            if(index==0):
                count+=1
                output_labels = np.genfromtxt(file_name_path, delimiter=',')
                b = sheet.cell(row=count, column=2)
                b.value = file_name
                for cols in range(output_labels.shape[1]):
                    ari=adjusted_rand_score(labels_true_gt,output_labels[:,cols])
                    print(ari)
                    c = sheet.cell(row=count, column=(cols+12))
                    c.value = ari
          except ValueError:
                print(
                "Not found!")
          else:
                print(
                "Found!")
print(count)
wb.save("C:/Users/dey.sn/Downloads/work/output/haraal_k6/results_python_ari_all_runs.xlsx")
