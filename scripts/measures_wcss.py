import pandas as pd
import numpy as np
#from scipy.spatial import distance
from math import dist
import os
import csv
import openpyxl
from sklearn.metrics.cluster import adjusted_rand_score

# https://github.com/cran/dendextend/blob/master/R/find_k.R
# https://cran.r-project.org/web/packages/fpc/fpc.pdf

# scipy.spatial.distance.euclidean(A, B)
# dist([1, 0, 0], [0, 1, 0])

data=np.genfromtxt("C:/Users/dey.sn/Downloads/temp/haraal/2d.csv", delimiter=',')
print(data.shape[0])
#print(data[10298])
vectors=data.shape[0]

# This is the path where you want to search
path = r'C:/Users/dey.sn/Downloads/work/output/haraal_k6/'
# this is the extension you want to detect
extension = '.csv'
substring="haraal_k6"
count=0
wb=openpyxl.Workbook()
sheet=wb.active
sheet.title= 'haraal'
for root, dirs_list, files_list in os.walk(path):
    for file_name in files_list:
       if os.path.splitext(file_name)[-1] == extension:
          file_name_path = os.path.join(root, file_name)
          print(file_name)
          print(file_name_path)   # This is the full path of the filter file
          try:
            index=file_name.index(substring)
 #          print(index)
            if(index==0):
                count+=1
                centarr = np.genfromtxt(file_name_path, delimiter=',')
                b = sheet.cell(row=count, column=2)
                b.value = file_name
#               centarr = np.genfromtxt('C:/Users/dey.sn/Downloads/work/output/har_k6/har_k6_kmeans_120cutoff _4_2.csv', delimiter=',')
#               print(np.shape(centarr))
#               print(centarr[0],centarr[1])
                index = 2
                row=int(centarr[0])   # number of centroids
                col=int(centarr[1])
                cents=[]
                for i in range(row):
                    c1=[]
                    for j in range(col):
                        c1.append(centarr[index])
                        index += 1
                    cents.append(c1)

#                print(cents[2])
#                print(np.shape(cents))

                wcss1=0
                for i in range (vectors):
                    distance1 = []
                    for j in range(row):
#                        print(j)
                        d1=(dist(data[i], cents[j]))
                        #print(d1)
                        distance1.append(d1)

                    print(distance1)
                    mindist=min(distance1)
                    print(mindist)

                    wcss1= int(wcss1 + (mindist*mindist))

                print("wcss1 is : " , (wcss1))

                c = sheet.cell(row=count, column=12)
                c.value = wcss1
          except ValueError:
                print
                "Not found!"
          else:
                print
                "Found!"
print(count)
wb.save("C:/Users/dey.sn/Downloads/work/output/haraal_k6/results_python_wcss_all_runs.xlsx")
